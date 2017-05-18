from Email_package.asegurado import *
from Email_package.ejecutivo import *
from Email_package.emailPropio import *
from Email_package.mensaje import *
import openpyxl

asunto = u'Vencimiento de póliza'

doc = openpyxl.load_workbook('C:\\Users\\Public\\Documents\\formato.xlsx')
# print(doc.get_sheet_names())
hoja = doc.get_sheet_by_name('Renovaciones')

emails  = []

f = 1



for fila in hoja.rows:

    dni = hoja.cell(row=f, column=5).value
    nombreAsegurado = hoja.cell(row=f, column=4).value
    if nombreAsegurado == None : break
    hasta = hoja.cell(row=f, column=10).value
    encargado = hoja.cell(row=f, column=11).value
    para = hoja.cell(row=f, column=13).value
    copiac = hoja.cell(row=f, column=14).value
    #print(f,dni,nombreAsegurado,hasta,encargado,para,copiac)

    f += 1
    if f == 2: continue

    asegurado = Asegurado(dni, nombreAsegurado, hasta)
    ejecutivo = Ejecutivo(encargado,para,copiac,asegurado)

    flag = False
    for email in emails:
        if email.TEXT.para == ejecutivo.nombre:
            flag = True
            email.TEXT.agregarLinea(ejecutivo)
            break

    print (flag)

    if flag == False :
        mensaje = Mensaje(ejecutivo)
        email = EmailPropio(u'byron.povea@outlook.es',u'miplumalamato06' , ejecutivo, asunto, mensaje)
        emails.append(email)


for email in emails:
    email.verEmail()
    #email.enviarEmail()


